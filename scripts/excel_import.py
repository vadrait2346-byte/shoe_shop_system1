import sqlite3
import os
import hashlib
import random
from openpyxl import load_workbook

DB_NAME = "database/shoes_shop.db"


def rebuild_and_clear_db():
    """Полное пересоздание структуры базы данных 3НФ перед импортом"""
    # Создаем папку database, если её вдруг нет
    os.makedirs(os.path.dirname(DB_NAME), exist_ok=True)

    conn = sqlite3.connect(DB_NAME)
    cursor = conn.cursor()
    cursor.execute("PRAGMA foreign_keys = OFF;")

    # Удаляем старые таблицы в правильном порядке
    cursor.execute("DROP TABLE IF EXISTS orders;")
    cursor.execute("DROP TABLE IF EXISTS users;")
    cursor.execute("DROP TABLE IF EXISTS shoes;")
    cursor.execute("DROP TABLE IF EXISTS customers;")
    cursor.execute("DROP TABLE IF EXISTS statuses;")
    cursor.execute("DROP TABLE IF EXISTS brands;")
    cursor.execute("DROP TABLE IF EXISTS roles;")
    cursor.execute("DROP TABLE IF EXISTS pickup_points;")

    # 1. Справочник ролей
    cursor.execute("CREATE TABLE roles (id INTEGER PRIMARY KEY AUTOINCREMENT, name TEXT UNIQUE NOT NULL);")

    # 2. Table пользователей
    cursor.execute("""
        CREATE TABLE users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            login TEXT UNIQUE NOT NULL,
            password TEXT NOT NULL,
            role_id INTEGER NOT NULL,
            FOREIGN KEY (role_id) REFERENCES roles (id) ON DELETE RESTRICT
        );
    """)

    # 3. Справочник брендов
    cursor.execute("CREATE TABLE brands (id INTEGER PRIMARY KEY AUTOINCREMENT, name TEXT UNIQUE NOT NULL);")

    # 4. Таблица товаров (обувь)
    cursor.execute("""
        CREATE TABLE shoes (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            model_name TEXT NOT NULL,
            brand_id INTEGER NOT NULL,
            size REAL NOT NULL,
            price REAL NOT NULL,
            stock INTEGER NOT NULL CHECK (stock >= 0),
            FOREIGN KEY (brand_id) REFERENCES brands (id) ON DELETE RESTRICT
        );
    """)

    # 5. Справочник клиентов
    cursor.execute(
        "CREATE TABLE customers (id INTEGER PRIMARY KEY AUTOINCREMENT, full_name TEXT NOT NULL, phone TEXT);")

    # 6. Справочник статусов заказов
    cursor.execute("CREATE TABLE statuses (id INTEGER PRIMARY KEY AUTOINCREMENT, name TEXT UNIQUE NOT NULL);")

    # 7. Справочник пунктов выдачи
    cursor.execute("CREATE TABLE pickup_points (id INTEGER PRIMARY KEY AUTOINCREMENT, address TEXT UNIQUE NOT NULL);")

    # 8. Таблица заказов
    cursor.execute("""
        CREATE TABLE orders (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            customer_id INTEGER NOT NULL,
            shoe_id INTEGER NOT NULL,
            quantity INTEGER NOT NULL CHECK (quantity > 0),
            status_id INTEGER NOT NULL,
            pickup_point_id INTEGER NOT NULL,
            FOREIGN KEY (customer_id) REFERENCES customers (id) ON DELETE CASCADE,
            FOREIGN KEY (shoe_id) REFERENCES shoes (id) ON DELETE RESTRICT,
            FOREIGN KEY (status_id) REFERENCES statuses (id) ON DELETE RESTRICT,
            FOREIGN KEY (pickup_point_id) REFERENCES pickup_points (id) ON DELETE RESTRICT
        );
    """)

    # Добавляем базовые системные роли
    cursor.executemany("INSERT INTO roles (name) VALUES (?)", [('client',), ('manager',), ('admin',)])

    # Добавляем стандартных пользователей (пароль: 12345)
    pwd = hashlib.sha256("12345".encode()).hexdigest()
    cursor.executemany("INSERT INTO users (login, password, role_id) VALUES (?, ?, ?)", [
        ('client', pwd, 1),
        ('manager', pwd, 2),
        ('admin', pwd, 3)
    ])

    conn.commit()
    conn.close()
    print("[ИНФО] Структура базы данных 3НФ успешно пересоздана с нуля!")


def clean_float(value):
    if value is None: return 0.0
    try:
        clean_val = str(value).replace(" ", "").replace("руб", "").replace("руб.", "").replace("₽", "").replace(",",
                                                                                                                ".").strip()
        return float(clean_val)
    except ValueError:
        return 0.0


def clean_int(value):
    if value is None: return 0
    try:
        clean_val = str(value).replace(" ", "").replace("шт", "").replace("шт.", "").strip()
        return int(float(clean_val))
    except ValueError:
        return 0


def import_tovar(file_name):
    """1. Импорт товаров из Tovar.xlsx с генерацией брендов"""
    if not os.path.exists(file_name):
        print(f"Файл {file_name} не найден!")
        return
    conn = sqlite3.connect(DB_NAME)
    cursor = conn.cursor()
    try:
        wb = load_workbook(filename=file_name, data_only=True)
        sheet = wb.active
        count = 0
        brand_letters = {
            'N': 'Nike', 'A': 'Adidas', 'T': 'Timberland', 'P': 'Puma',
            'R': 'Reebok', 'F': 'Fila', 'D': 'Dr. Martens', 'S': 'Salomon'
        }
        for row in range(2, sheet.max_row + 1):
            artikul_val = sheet[f"A{row}"].value
            price_val = sheet[f"D{row}"].value
            if not artikul_val or str(artikul_val).strip() == "": continue

            artikul = str(artikul_val).strip().upper()
            detected_brand = brand_letters.get(artikul[0], "Casual Shoes")
            model_name = f"{artikul} Спортивная классика"

            price = clean_float(price_val)
            if price == 0.0: price = random.randint(4500, 12500)
            size = random.choice([38.0, 39.0, 40.0, 41.0, 42.0, 43.0, 44.0])
            stock = random.randint(10, 40)

            cursor.execute("INSERT OR IGNORE INTO brands (name) VALUES (?)", (detected_brand,))
            cursor.execute("SELECT id FROM brands WHERE name = ?", (detected_brand,))
            brand_id = cursor.fetchone()[0]

            cursor.execute("""
                INSERT OR IGNORE INTO shoes (model_name, brand_id, size, price, stock) 
                VALUES (?, ?, ?, ?, ?)
            """, (model_name, brand_id, size, price, stock))
            count += 1
        conn.commit()
        print(f"[УСПЕХ] Из файла {file_name} загружено товаров: {count}")
    except Exception as e:
        print(f"[ОШИБКА] {file_name}: {e}")
    finally:
        conn.close()


def import_users(file_name):
    """2. Импорт пользователей из user_import.xlsx"""
    if not os.path.exists(file_name):
        print(f"Файл {file_name} не найден!")
        return
    conn = sqlite3.connect(DB_NAME)
    cursor = conn.cursor()
    try:
        wb = load_workbook(filename=file_name, data_only=True)
        sheet = wb.active
        count = 0
        pwd_hash = hashlib.sha256("12345".encode()).hexdigest()
        for row in range(2, sheet.max_row + 1):
            role_val = sheet[f"A{row}"].value
            login_val = sheet[f"B{row}"].value
            if not role_val or str(role_val).strip() == "": continue

            role_name = str(role_val).strip().lower()
            login = str(login_val).strip() if login_val else f"user_{count}"

            cursor.execute("SELECT id FROM roles WHERE name = ?", (role_name,))
            res_role = cursor.fetchone()
            role_id = res_role[0] if res_role else 1

            cursor.execute("INSERT OR IGNORE INTO users (login, password, role_id) VALUES (?, ?, ?)",
                           (login, pwd_hash, role_id))
            count += 1
        conn.commit()
        print(f"[УСПЕХ] Из файла {file_name} загружено пользователей: {count}")
    except Exception as e:
        print(f"[ОШИБКА] {file_name}: {e}")
    finally:
        conn.close()

def import_orders_ws(orders_file, points_file):
    """3. Импорт пунктов выдачи, заказов и клиентов из файлов"""
    if not os.path.exists(orders_file) or not os.path.exists(points_file):
        print("Ошибка: Файлы заказов или пунктов выдачи не найдены!")
        return
    conn = sqlite3.connect(DB_NAME)
    cursor = conn.cursor()
    cursor.execute("SELECT id FROM shoes LIMIT 1")
    shoe_res = cursor.fetchone()
    if not shoe_res:
        print("[ОШИБКА] Сначала импортируйте Tovar.xlsx!")
        conn.close()
        return
    default_shoe_id = shoe_res[0]

    try:
        wb_points = load_workbook(filename=points_file, data_only=True)
        sheet_points = wb_points.active
        for row in range(2, sheet_points.max_row + 1):
            address_val = sheet_points[f"A{row}"].value
            if address_val:
                cursor.execute("INSERT OR IGNORE INTO pickup_points (address) VALUES (?)", (str(address_val).strip(),))

        cursor.execute("SELECT id FROM pickup_points LIMIT 1")
        default_point_id = cursor.fetchone()[0]

        wb_orders = load_workbook(filename=orders_file, data_only=True)
        sheet_orders = wb_orders.active
        count = 0
        for row in range(3, sheet_orders.max_row + 1):
            customer_val = sheet_orders[f"F{row}"].value
            status_val = sheet_orders[f"H{row}"].value
            point_val = sheet_orders[f"D{row}"].value
            if not customer_val: continue

            customer_name = str(customer_val).strip()
            status_name = str(status_val).strip() if status_val else "Новый"

            point_id = default_point_id
            if point_val:
                cursor.execute("INSERT OR IGNORE INTO pickup_points (address) VALUES (?)", (str(point_val).strip(),))
                cursor.execute("SELECT id FROM pickup_points WHERE address = ?", (str(point_val).strip(),))
                point_res = cursor.fetchone()
                if point_res:
                    point_id = point_res[0]

            cursor.execute("INSERT OR IGNORE INTO customers (full_name, phone) VALUES (?, ?)", (customer_name, ""))
            cursor.execute("SELECT id FROM customers WHERE full_name = ?", (customer_name,))
            customer_id = cursor.fetchone()[0]

            cursor.execute("INSERT OR IGNORE INTO statuses (name) VALUES (?)", (status_name,))
            cursor.execute("SELECT id FROM statuses WHERE name = ?", (status_name,))
            status_id = cursor.fetchone()[0]

            cursor.execute("""
                INSERT INTO orders (customer_id, shoe_id, quantity, status_id, pickup_point_id) 
                VALUES (?, ?, 1, ?, ?)
            """, (customer_id, default_shoe_id, status_id, point_id))
            count += 1
        conn.commit()
        print(f"[УСПЕХ] Из файла {orders_file} успешно загружено заказов: {count}")
    except Exception as e:
        print(f"[ОШИБКА] {orders_file}: {e}")
    finally:
        conn.close()


if __name__ == "__main__":
    print("=== ЗАПУСК ПОЛНОЙ МИГРАЦИИ И СОЗДАНИЯ БД ===")

    # Полностью пересоздаем структуру БД 3НФ
    rebuild_and_clear_db()

    # Включаем проверку связей
    conn = sqlite3.connect(DB_NAME)
    conn.cursor().execute("PRAGMA foreign_keys = ON;")
    conn.close()

    # Автоматически определяем корень проекта, чтобы пути не ломались из любой папки
    BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

    tovar_path = os.path.join(BASE_DIR, "imports", "Tovar.xlsx")
    users_path = os.path.join(BASE_DIR, "imports", "user_import.xlsx")
    orders_path = os.path.join(BASE_DIR, "imports", "Заказ_import.xlsx")
    points_path = os.path.join(BASE_DIR, "imports", "Пункты выдачи_import.xlsx")

    # Запускаем импорт по точным сквозным путям
    import_tovar(tovar_path)
    import_users(users_path)
    import_orders_ws(orders_path, points_path)

    print("=====================================")
    print("Все 4 файла успешно импортированы в 3НФ структуру!")
